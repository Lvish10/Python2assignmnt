import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
import os
import bcrypt
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from PIL import Image, ImageTk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook

# Get the directory of the current script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# CSV file paths (relative to the script directory)
STUDENT_FILE = os.path.join(SCRIPT_DIR, 'students.csv')
MARKS_FILE = os.path.join(SCRIPT_DIR, 'marks.csv')
ATTENDANCE_FILE = os.path.join(SCRIPT_DIR, 'attendance.csv')
MODULES_FILE = os.path.join(SCRIPT_DIR, 'modules.csv')
QUIZZES_FILE = os.path.join(SCRIPT_DIR, 'quizzes.csv')
NOTES_FILE = os.path.join(SCRIPT_DIR, 'notes.csv')
DISCUSSION_FILE = os.path.join(SCRIPT_DIR, 'discussion.csv')

# Initialize CSV files with headers if they don't exist
def initialize_files():
    if not os.path.exists(STUDENT_FILE):
        with open(STUDENT_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Student ID', 'Name', 'Age', 'Class', 'Address', 'Medical Info', 'Password', 'Cohort'])

    if not os.path.exists(MARKS_FILE):
        with open(MARKS_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Student ID', 'Subject', 'Marks', 'Comments'])

    if not os.path.exists(ATTENDANCE_FILE):
        with open(ATTENDANCE_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Student ID', 'Subject', 'Attendance'])

    if not os.path.exists(MODULES_FILE):
        with open(MODULES_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Student ID', 'Module', 'Subject'])

    if not os.path.exists(QUIZZES_FILE):
        with open(QUIZZES_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Module', 'Quiz', 'Answer'])

    if not os.path.exists(NOTES_FILE):
        with open(NOTES_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Module', 'Note'])

    if not os.path.exists(DISCUSSION_FILE):
        with open(DISCUSSION_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Module', 'User', 'Message'])

initialize_files()

# Function to read data from CSV
def read_csv(file):
    with open(file, mode='r') as f:
        return list(csv.reader(f))

# Function to write data to CSV
def write_csv(file, data):
    with open(file, mode='a', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(data)

# Function to update data in CSV
def update_csv(file, data):
    with open(file, mode='w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Hash password using bcrypt
def hash_password(password):
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt())

# Verify password using bcrypt
def verify_password(password, hashed):
    return bcrypt.checkpw(password.encode(), hashed.encode())

# Send email
def send_email(to_email, subject, body):
    from_email = "your_email@example.com"  # Replace with your email
    password = "your_password"  # Replace with your email password

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_email, password)
        server.sendmail(from_email, to_email, msg.as_string())
        server.quit()
        messagebox.showinfo("Success", "Email sent successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to send email: {e}")

# Export data to Excel
def export_to_excel():
    wb = Workbook()
    ws = wb.active

    # Export students
    ws.title = "Students"
    students = read_csv(STUDENT_FILE)
    for row in students:
        ws.append(row)

    # Export marks
    ws = wb.create_sheet("Marks")
    marks = read_csv(MARKS_FILE)
    for row in marks:
        ws.append(row)

    # Export attendance
    ws = wb.create_sheet("Attendance")
    attendance = read_csv(ATTENDANCE_FILE)
    for row in attendance:
        ws.append(row)

    # Export modules
    ws = wb.create_sheet("Modules")
    modules = read_csv(MODULES_FILE)
    for row in modules:
        ws.append(row)

    # Save file
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialdir=SCRIPT_DIR)
    if file_path:
        wb.save(file_path)
        messagebox.showinfo("Success", "Data exported successfully!")

# Teacher Interface
class TeacherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Teacher Interface")
        self.create_widgets()

    def create_widgets(self):
        # Notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.grid(row=0, column=0, padx=10, pady=10)

        # Add Student Tab
        self.add_student_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.add_student_tab, text="Add Student")
        self.create_add_student_tab()

        # View Students Tab
        self.view_students_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.view_students_tab, text="View Students")
        self.create_view_students_tab()

        # Manage Attendance Tab
        self.attendance_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.attendance_tab, text="Manage Attendance")
        self.create_attendance_tab()

        # Manage Modules Tab
        self.modules_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.modules_tab, text="Manage Modules")
        self.create_modules_tab()

        # Manage Quizzes Tab
        self.quizzes_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.quizzes_tab, text="Manage Quizzes")
        self.create_quizzes_tab()

        # Manage Notes Tab
        self.notes_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.notes_tab, text="Manage Notes")
        self.create_notes_tab()

        # Discussion Forum Tab
        self.discussion_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.discussion_tab, text="Discussion Forum")
        self.create_discussion_tab()

        # Report Cards Tab
        self.report_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.report_tab, text="Report Cards")
        self.create_report_tab()

        # Export Data Tab
        self.export_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.export_tab, text="Export Data")
        self.create_export_tab()

        # Visualizations Tab
        self.visualizations_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.visualizations_tab, text="Visualizations")
        self.create_visualizations_tab()

    def create_add_student_tab(self):
        add_frame = ttk.LabelFrame(self.add_student_tab, text="Add Student")
        add_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Label(add_frame, text="Student ID:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_id = ttk.Entry(add_frame)
        self.entry_id.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(add_frame, text="Name:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_name = ttk.Entry(add_frame)
        self.entry_name.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(add_frame, text="Age:").grid(row=2, column=0, padx=5, pady=5)
        self.entry_age = ttk.Entry(add_frame)
        self.entry_age.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(add_frame, text="Class:").grid(row=3, column=0, padx=5, pady=5)
        self.entry_class = ttk.Entry(add_frame)
        self.entry_class.grid(row=3, column=1, padx=5, pady=5)

        ttk.Label(add_frame, text="Address:").grid(row=4, column=0, padx=5, pady=5)
        self.entry_address = ttk.Entry(add_frame)
        self.entry_address.grid(row=4, column=1, padx=5, pady=5)

        ttk.Label(add_frame, text="Medical Info:").grid(row=5, column=0, padx=5, pady=5)
        self.entry_medical = ttk.Entry(add_frame)
        self.entry_medical.grid(row=5, column=1, padx=5, pady=5)

        ttk.Label(add_frame, text="Password:").grid(row=6, column=0, padx=5, pady=5)
        self.entry_password = ttk.Entry(add_frame, show='*')
        self.entry_password.grid(row=6, column=1, padx=5, pady=5)

        ttk.Label(add_frame, text="Cohort:").grid(row=7, column=0, padx=5, pady=5)
        self.cohort_var = tk.StringVar()
        self.cohort_dropdown = ttk.Combobox(add_frame, textvariable=self.cohort_var)
        self.cohort_dropdown['values'] = [
            "Diploma in Emerging Tech (IoT) Cohort 6",
            "Diploma in Emerging Tech (BDA) Cohort 6",
            "Diploma in Cybersecurity Cohort 5",
            "Diploma in AI and ML Cohort 4",
            "Diploma in Cloud Computing Cohort 3",
            "Diploma in Software Engineering Cohort 7",
            "Diploma in Data Science Cohort 5",
            "Diploma in Blockchain Technology Cohort 2",
            "Diploma in Full Stack Development Cohort 8",
            "Diploma in DevOps Cohort 4",
            "Diploma in Mobile App Development Cohort 6",
            "Diploma in Game Development Cohort 3",
            "Diploma in UI/UX Design Cohort 5",
            "Diploma in IT Project Management Cohort 4",
            "Diploma in Network Engineering Cohort 6",
            "Diploma in Digital Marketing Cohort 7",
            "Diploma in IT Support and Services Cohort 5",
            "Diploma in Big Data Analytics Cohort 4",
            "Diploma in Machine Learning Cohort 3",
            "Diploma in Quantum Computing Cohort 1"
        ]
        self.cohort_dropdown.grid(row=7, column=1, padx=5, pady=5)

        ttk.Button(add_frame, text="Add Student", command=self.add_student).grid(row=8, column=0, columnspan=2, pady=10)

    def create_view_students_tab(self):
        view_frame = ttk.LabelFrame(self.view_students_tab, text="View Students")
        view_frame.grid(row=0, column=0, padx=10, pady=10)

        self.tree = ttk.Treeview(view_frame, columns=('ID', 'Name', 'Age', 'Class', 'Address', 'Medical Info', 'Cohort'), show='headings')
        self.tree.heading('ID', text='ID')
        self.tree.heading('Name', text='Name')
        self.tree.heading('Age', text='Age')
        self.tree.heading('Class', text='Class')
        self.tree.heading('Address', text='Address')
        self.tree.heading('Medical Info', text='Medical Info')
        self.tree.heading('Cohort', text='Cohort')
        self.tree.grid(row=0, column=0, padx=5, pady=5)

        ttk.Button(view_frame, text="Refresh", command=self.refresh_students).grid(row=1, column=0, pady=5)

        self.refresh_students()

    def create_attendance_tab(self):
        attendance_frame = ttk.LabelFrame(self.attendance_tab, text="Manage Attendance")
        attendance_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Label(attendance_frame, text="Student ID:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_attendance_id = ttk.Entry(attendance_frame)
        self.entry_attendance_id.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(attendance_frame, text="Subject:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_attendance_subject = ttk.Entry(attendance_frame)
        self.entry_attendance_subject.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(attendance_frame, text="Attendance:").grid(row=2, column=0, padx=5, pady=5)
        self.entry_attendance = ttk.Entry(attendance_frame)
        self.entry_attendance.grid(row=2, column=1, padx=5, pady=5)

        ttk.Button(attendance_frame, text="Mark Attendance", command=self.mark_attendance).grid(row=3, column=0, columnspan=2, pady=10)

    def create_modules_tab(self):
        modules_frame = ttk.LabelFrame(self.modules_tab, text="Manage Modules")
        modules_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Label(modules_frame, text="Student ID:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_module_id = ttk.Entry(modules_frame)
        self.entry_module_id.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(modules_frame, text="Module:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_module = ttk.Entry(modules_frame)
        self.entry_module.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(modules_frame, text="Subject:").grid(row=2, column=0, padx=5, pady=5)
        self.entry_module_subject = ttk.Entry(modules_frame)
        self.entry_module_subject.grid(row=2, column=1, padx=5, pady=5)

        ttk.Button(modules_frame, text="Add Module", command=self.add_module).grid(row=3, column=0, columnspan=2, pady=10)

    def create_quizzes_tab(self):
        quizzes_frame = ttk.LabelFrame(self.quizzes_tab, text="Manage Quizzes")
        quizzes_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Label(quizzes_frame, text="Module:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_quiz_module = ttk.Entry(quizzes_frame)
        self.entry_quiz_module.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(quizzes_frame, text="Quiz:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_quiz = ttk.Entry(quizzes_frame)
        self.entry_quiz.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(quizzes_frame, text="Answer:").grid(row=2, column=0, padx=5, pady=5)
        self.entry_quiz_answer = ttk.Entry(quizzes_frame)
        self.entry_quiz_answer.grid(row=2, column=1, padx=5, pady=5)

        ttk.Button(quizzes_frame, text="Add Quiz", command=self.add_quiz).grid(row=3, column=0, columnspan=2, pady=10)

    def create_notes_tab(self):
        notes_frame = ttk.LabelFrame(self.notes_tab, text="Manage Notes")
        notes_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Label(notes_frame, text="Module:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_note_module = ttk.Entry(notes_frame)
        self.entry_note_module.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(notes_frame, text="Note:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_note = ttk.Entry(notes_frame)
        self.entry_note.grid(row=1, column=1, padx=5, pady=5)

        ttk.Button(notes_frame, text="Add Note", command=self.add_note).grid(row=2, column=0, columnspan=2, pady=10)

    def create_discussion_tab(self):
        discussion_frame = ttk.LabelFrame(self.discussion_tab, text="Discussion Forum")
        discussion_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Label(discussion_frame, text="Module:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_discussion_module = ttk.Entry(discussion_frame)
        self.entry_discussion_module.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(discussion_frame, text="Message:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_discussion_message = ttk.Entry(discussion_frame)
        self.entry_discussion_message.grid(row=1, column=1, padx=5, pady=5)

        ttk.Button(discussion_frame, text="Post Message", command=self.post_message).grid(row=2, column=0, columnspan=2, pady=10)

    def create_report_tab(self):
        report_frame = ttk.LabelFrame(self.report_tab, text="Generate Report Card")
        report_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Label(report_frame, text="Student ID:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_report_id = ttk.Entry(report_frame)
        self.entry_report_id.grid(row=0, column=1, padx=5, pady=5)

        ttk.Button(report_frame, text="Generate Report", command=self.generate_report).grid(row=1, column=0, columnspan=2, pady=10)

    def create_export_tab(self):
        export_frame = ttk.LabelFrame(self.export_tab, text="Export Data")
        export_frame.grid(row=0, column=0, padx=10, pady=10)

        ttk.Button(export_frame, text="Export to Excel", command=export_to_excel).grid(row=0, column=0, pady=10)

    def create_visualizations_tab(self):
        # Frame for visualizations
        visualizations_frame = ttk.LabelFrame(self.visualizations_tab, text="Visualizations")
        visualizations_frame.grid(row=0, column=0, padx=10, pady=10)

        # Buttons to toggle visualizations
        self.attendance_chart_button = ttk.Button(visualizations_frame, text="Show Attendance Chart", command=lambda: self.toggle_chart(visualizations_frame, "attendance"))
        self.attendance_chart_button.grid(row=0, column=0, pady=10)

        self.marks_chart_button = ttk.Button(visualizations_frame, text="Show Marks Chart", command=lambda: self.toggle_chart(visualizations_frame, "marks"))
        self.marks_chart_button.grid(row=1, column=0, pady=10)

        self.modules_chart_button = ttk.Button(visualizations_frame, text="Show Modules Chart", command=lambda: self.toggle_chart(visualizations_frame, "modules"))
        self.modules_chart_button.grid(row=2, column=0, pady=10)

        self.attendance_trends_button = ttk.Button(visualizations_frame, text="Show Attendance Trends", command=lambda: self.toggle_chart(visualizations_frame, "attendance_trends"))
        self.attendance_trends_button.grid(row=3, column=0, pady=10)

        self.marks_distribution_button = ttk.Button(visualizations_frame, text="Show Marks Distribution", command=lambda: self.toggle_chart(visualizations_frame, "marks_distribution"))
        self.marks_distribution_button.grid(row=4, column=0, pady=10)

        # Dictionary to store chart widgets
        self.chart_widgets = {}

    def toggle_chart(self, frame, chart_type):
        if chart_type in self.chart_widgets:
            # If the chart is already displayed, hide it
            self.chart_widgets[chart_type].get_tk_widget().grid_forget()
            del self.chart_widgets[chart_type]
        else:
            # If the chart is not displayed, show it
            if chart_type == "attendance":
                self.show_attendance_chart(frame)
            elif chart_type == "marks":
                self.show_marks_chart(frame)
            elif chart_type == "modules":
                self.show_modules_chart(frame)
            elif chart_type == "attendance_trends":
                self.show_attendance_trends(frame)
            elif chart_type == "marks_distribution":
                self.show_marks_distribution(frame)

    def show_attendance_chart(self, frame):
        attendance_data = read_csv(ATTENDANCE_FILE)
        subjects = {}
        for row in attendance_data[1:]:
            subject = row[1]
            attendance = int(row[2])
            if subject in subjects:
                subjects[subject].append(attendance)
            else:
                subjects[subject] = [attendance]

        avg_attendance = {subject: sum(attendance_list) / len(attendance_list) for subject, attendance_list in subjects.items()}

        fig, ax = plt.subplots()
        ax.bar(avg_attendance.keys(), avg_attendance.values())
        ax.set_xlabel('Subject')
        ax.set_ylabel('Average Attendance (%)')
        ax.set_title('Average Attendance by Subject')

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().grid(row=5, column=0, padx=10, pady=10)
        self.chart_widgets["attendance"] = canvas

    def show_marks_chart(self, frame):
        marks_data = read_csv(MARKS_FILE)
        subjects = {}
        for row in marks_data[1:]:
            subject = row[1]
            marks = int(row[2])
            if subject in subjects:
                subjects[subject].append(marks)
            else:
                subjects[subject] = [marks]

        avg_marks = {subject: sum(marks_list) / len(marks_list) for subject, marks_list in subjects.items()}

        fig, ax = plt.subplots()
        ax.bar(avg_marks.keys(), avg_marks.values())
        ax.set_xlabel('Subject')
        ax.set_ylabel('Average Marks')
        ax.set_title('Average Marks by Subject')

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().grid(row=6, column=0, padx=10, pady=10)
        self.chart_widgets["marks"] = canvas

    def show_modules_chart(self, frame):
        modules_data = read_csv(MODULES_FILE)
        modules = {}
        for row in modules_data[1:]:
            module = row[1]
            if module in modules:
                modules[module] += 1
            else:
                modules[module] = 1

        fig, ax = plt.subplots()
        ax.pie(modules.values(), labels=modules.keys(), autopct='%1.1f%%')
        ax.set_title('Module Distribution')

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().grid(row=7, column=0, padx=10, pady=10)
        self.chart_widgets["modules"] = canvas

    def show_attendance_trends(self, frame):
        attendance_data = read_csv(ATTENDANCE_FILE)
        dates = [f"Day {i+1}" for i in range(len(attendance_data[1:]))]
        attendance = [int(row[2]) for row in attendance_data[1:]]

        fig, ax = plt.subplots()
        ax.plot(dates, attendance, marker='o')
        ax.set_xlabel('Date')
        ax.set_ylabel('Attendance (%)')
        ax.set_title('Attendance Trends Over Time')

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().grid(row=8, column=0, padx=10, pady=10)
        self.chart_widgets["attendance_trends"] = canvas

    def show_marks_distribution(self, frame):
        marks_data = read_csv(MARKS_FILE)
        marks = [int(row[2]) for row in marks_data[1:]]

        fig, ax = plt.subplots()
        ax.hist(marks, bins=10, edgecolor='black')
        ax.set_xlabel('Marks')
        ax.set_ylabel('Frequency')
        ax.set_title('Marks Distribution')

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().grid(row=9, column=0, padx=10, pady=10)
        self.chart_widgets["marks_distribution"] = canvas

    def add_student(self):
        student_id = self.entry_id.get()
        name = self.entry_name.get()
        age = self.entry_age.get()
        student_class = self.entry_class.get()
        address = self.entry_address.get()
        medical_info = self.entry_medical.get()
        password = self.entry_password.get()
        cohort = self.cohort_var.get()

        if student_id and name and age and student_class and address and medical_info and password and cohort:
            hashed_password = hash_password(password)
            write_csv(STUDENT_FILE, [student_id, name, age, student_class, address, medical_info, hashed_password, cohort])
            messagebox.showinfo("Success", "Student added successfully!")
            self.refresh_students()
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")

    def refresh_students(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        students = read_csv(STUDENT_FILE)
        for student in students[1:]:
            if len(student) >= 8:
                self.tree.insert('', 'end', values=(student[0], student[1], student[2], student[3], student[4], student[5], student[7]))
            else:
                print(f"Skipping invalid row: {student}")

    def mark_attendance(self):
        student_id = self.entry_attendance_id.get()
        subject = self.entry_attendance_subject.get()
        attendance = self.entry_attendance.get()

        if student_id and subject and attendance:
            write_csv(ATTENDANCE_FILE, [student_id, subject, attendance])
            messagebox.showinfo("Success", "Attendance marked successfully!")
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")

    def add_module(self):
        student_id = self.entry_module_id.get()
        module = self.entry_module.get()
        subject = self.entry_module_subject.get()

        if student_id and module and subject:
            write_csv(MODULES_FILE, [student_id, module, subject])
            messagebox.showinfo("Success", "Module added successfully!")
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")

    def add_quiz(self):
        module = self.entry_quiz_module.get()
        quiz = self.entry_quiz.get()
        answer = self.entry_quiz_answer.get()

        if module and quiz and answer:
            write_csv(QUIZZES_FILE, [module, quiz, answer])
            messagebox.showinfo("Success", "Quiz added successfully!")
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")

    def add_note(self):
        module = self.entry_note_module.get()
        note = self.entry_note.get()

        if module and note:
            write_csv(NOTES_FILE, [module, note])
            messagebox.showinfo("Success", "Note added successfully!")
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")

    def post_message(self):
        module = self.entry_discussion_module.get()
        message = self.entry_discussion_message.get()

        if module and message:
            write_csv(DISCUSSION_FILE, [module, "Teacher", message])
            messagebox.showinfo("Success", "Message posted successfully!")
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")

    def generate_report(self):
        student_id = self.entry_report_id.get()
        if student_id:
            students = read_csv(STUDENT_FILE)
            marks = read_csv(MARKS_FILE)
            attendance = read_csv(ATTENDANCE_FILE)
            modules = read_csv(MODULES_FILE)

            student_info = next((student for student in students[1:] if student[0] == student_id), None)
            if student_info:
                report = f"Report Card for {student_info[1]} (ID: {student_info[0]})\n\n"
                report += f"Class: {student_info[3]}\n"
                report += f"Age: {student_info[2]}\n"
                report += f"Address: {student_info[4]}\n"
                report += f"Medical Info: {student_info[5]}\n\n"

                report += "Marks:\n"
                for mark in marks[1:]:
                    if mark[0] == student_id:
                        report += f"{mark[1]}: {mark[2]} ({mark[3]})\n"

                report += "\nAttendance:\n"
                for att in attendance[1:]:
                    if att[0] == student_id:
                        report += f"{att[1]}: {att[2]}%\n"

                report += "\nModules:\n"
                for module in modules[1:]:
                    if module[0] == student_id:
                        report += f"{module[1]} ({module[2]})\n"

                messagebox.showinfo("Report Card", report)
            else:
                messagebox.showwarning("Not Found", "Student ID not found")
        else:
            messagebox.showwarning("Input Error", "Please enter a Student ID")

# Student Interface
class StudentApp:
    def __init__(self, root, student_id):
        self.root = root
        self.root.title("Student Interface")
        self.student_id = student_id
        self.create_widgets()

    def create_widgets(self):
        # Notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.grid(row=0, column=0, padx=10, pady=10)

        # Student Info Tab
        self.info_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.info_tab, text="Student Info")
        self.create_info_tab()

        # Marks Tab
        self.marks_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.marks_tab, text="Marks")
        self.create_marks_tab()

        # Attendance Tab
        self.attendance_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.attendance_tab, text="Attendance")
        self.create_attendance_tab()

        # Modules Tab
        self.modules_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.modules_tab, text="Modules")
        self.create_modules_tab()

        # Quizzes Tab
        self.quizzes_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.quizzes_tab, text="Quizzes")
        self.create_quizzes_tab()

        # Notes Tab
        self.notes_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.notes_tab, text="Notes")
        self.create_notes_tab()

        # Discussion Tab
        self.discussion_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.discussion_tab, text="Discussion")
        self.create_discussion_tab()

    def create_info_tab(self):
        info_frame = ttk.LabelFrame(self.info_tab, text="Student Info")
        info_frame.grid(row=0, column=0, padx=10, pady=10)

        students = read_csv(STUDENT_FILE)
        for student in students[1:]:
            if student[0] == self.student_id:
                ttk.Label(info_frame, text=f"ID: {student[0]}").grid(row=0, column=0, padx=5, pady=5)
                ttk.Label(info_frame, text=f"Name: {student[1]}").grid(row=1, column=0, padx=5, pady=5)
                ttk.Label(info_frame, text=f"Age: {student[2]}").grid(row=2, column=0, padx=5, pady=5)
                ttk.Label(info_frame, text=f"Class: {student[3]}").grid(row=3, column=0, padx=5, pady=5)
                ttk.Label(info_frame, text=f"Address: {student[4]}").grid(row=4, column=0, padx=5, pady=5)
                ttk.Label(info_frame, text=f"Medical Info: {student[5]}").grid(row=5, column=0, padx=5, pady=5)
                break

    def create_marks_tab(self):
        marks_frame = ttk.LabelFrame(self.marks_tab, text="Marks")
        marks_frame.grid(row=0, column=0, padx=10, pady=10)

        self.tree_marks = ttk.Treeview(marks_frame, columns=('Subject', 'Marks', 'Comments'), show='headings')
        self.tree_marks.heading('Subject', text='Subject')
        self.tree_marks.heading('Marks', text='Marks')
        self.tree_marks.heading('Comments', text='Comments')
        self.tree_marks.grid(row=0, column=0, padx=5, pady=5)

        self.refresh_marks()

    def create_attendance_tab(self):
        attendance_frame = ttk.LabelFrame(self.attendance_tab, text="Attendance")
        attendance_frame.grid(row=0, column=0, padx=10, pady=10)

        self.tree_attendance = ttk.Treeview(attendance_frame, columns=('Subject', 'Attendance'), show='headings')
        self.tree_attendance.heading('Subject', text='Subject')
        self.tree_attendance.heading('Attendance', text='Attendance')
        self.tree_attendance.grid(row=0, column=0, padx=5, pady=5)

        self.refresh_attendance()

    def create_modules_tab(self):
        modules_frame = ttk.LabelFrame(self.modules_tab, text="Modules")
        modules_frame.grid(row=0, column=0, padx=10, pady=10)

        self.tree_modules = ttk.Treeview(modules_frame, columns=('Module', 'Subject'), show='headings')
        self.tree_modules.heading('Module', text='Module')
        self.tree_modules.heading('Subject', text='Subject')
        self.tree_modules.grid(row=0, column=0, padx=5, pady=5)

        self.refresh_modules()

    def create_quizzes_tab(self):
        quizzes_frame = ttk.LabelFrame(self.quizzes_tab, text="Quizzes")
        quizzes_frame.grid(row=0, column=0, padx=10, pady=10)

        self.tree_quizzes = ttk.Treeview(quizzes_frame, columns=('Module', 'Quiz', 'Answer'), show='headings')
        self.tree_quizzes.heading('Module', text='Module')
        self.tree_quizzes.heading('Quiz', text='Quiz')
        self.tree_quizzes.heading('Answer', text='Answer')
        self.tree_quizzes.grid(row=0, column=0, padx=5, pady=5)

        self.refresh_quizzes()

    def create_notes_tab(self):
        notes_frame = ttk.LabelFrame(self.notes_tab, text="Notes")
        notes_frame.grid(row=0, column=0, padx=10, pady=10)

        self.tree_notes = ttk.Treeview(notes_frame, columns=('Module', 'Note'), show='headings')
        self.tree_notes.heading('Module', text='Module')
        self.tree_notes.heading('Note', text='Note')
        self.tree_notes.grid(row=0, column=0, padx=5, pady=5)

        self.refresh_notes()

    def create_discussion_tab(self):
        discussion_frame = ttk.LabelFrame(self.discussion_tab, text="Discussion")
        discussion_frame.grid(row=0, column=0, padx=10, pady=10)

        self.tree_discussion = ttk.Treeview(discussion_frame, columns=('Module', 'User', 'Message'), show='headings')
        self.tree_discussion.heading('Module', text='Module')
        self.tree_discussion.heading('User', text='User')
        self.tree_discussion.heading('Message', text='Message')
        self.tree_discussion.grid(row=0, column=0, padx=5, pady=5)

        self.refresh_discussion()

    def refresh_marks(self):
        for row in self.tree_marks.get_children():
            self.tree_marks.delete(row)
        marks = read_csv(MARKS_FILE)
        for mark in marks[1:]:
            if mark[0] == self.student_id:
                self.tree_marks.insert('', 'end', values=mark[1:])

    def refresh_attendance(self):
        for row in self.tree_attendance.get_children():
            self.tree_attendance.delete(row)
        attendance = read_csv(ATTENDANCE_FILE)
        for att in attendance[1:]:
            if att[0] == self.student_id:
                self.tree_attendance.insert('', 'end', values=att[1:])

    def refresh_modules(self):
        for row in self.tree_modules.get_children():
            self.tree_modules.delete(row)
        modules = read_csv(MODULES_FILE)
        for module in modules[1:]:
            if module[0] == self.student_id:
                self.tree_modules.insert('', 'end', values=module[1:])

    def refresh_quizzes(self):
        for row in self.tree_quizzes.get_children():
            self.tree_quizzes.delete(row)
        quizzes = read_csv(QUIZZES_FILE)
        for quiz in quizzes[1:]:
            self.tree_quizzes.insert('', 'end', values=quiz)

    def refresh_notes(self):
        for row in self.tree_notes.get_children():
            self.tree_notes.delete(row)
        notes = read_csv(NOTES_FILE)
        for note in notes[1:]:
            self.tree_notes.insert('', 'end', values=note)

    def refresh_discussion(self):
        for row in self.tree_discussion.get_children():
            self.tree_discussion.delete(row)
        discussion = read_csv(DISCUSSION_FILE)
        for message in discussion[1:]:
            self.tree_discussion.insert('', 'end', values=message)

def verify_password(password, hashed):
    if isinstance(hashed, str):
        hashed = hashed[2:-1].encode()
    return bcrypt.checkpw(password.encode(), hashed)

# Login Interface
class LoginApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Login")
        self.create_widgets()

    def create_widgets(self):
        login_type_frame = ttk.LabelFrame(self.root, text="Login As")
        login_type_frame.grid(row=0, column=0, padx=10, pady=10)

        self.login_type = tk.StringVar(value="teacher")
        ttk.Radiobutton(login_type_frame, text="Teacher", variable=self.login_type, value="teacher", command=self.update_login_fields).grid(row=0, column=0, padx=5, pady=5)
        ttk.Radiobutton(login_type_frame, text="Student", variable=self.login_type, value="student", command=self.update_login_fields).grid(row=0, column=1, padx=5, pady=5)

        self.login_fields_frame = ttk.LabelFrame(self.root, text="Login Details")
        self.login_fields_frame.grid(row=1, column=0, padx=10, pady=10)

        self.label_username = ttk.Label(self.login_fields_frame, text="Username:")
        self.label_username.grid(row=0, column=0, padx=5, pady=5)
        self.entry_username = ttk.Entry(self.login_fields_frame)
        self.entry_username.grid(row=0, column=1, padx=5, pady=5)

        self.label_student_id = ttk.Label(self.login_fields_frame, text="Student ID:")
        self.label_student_id.grid(row=1, column=0, padx=5, pady=5)
        self.entry_student_id = ttk.Entry(self.login_fields_frame)
        self.entry_student_id.grid(row=1, column=1, padx=5, pady=5)

        self.label_password = ttk.Label(self.login_fields_frame, text="Password:")
        self.label_password.grid(row=2, column=0, padx=5, pady=5)
        self.entry_password = ttk.Entry(self.login_fields_frame, show='*')
        self.entry_password.grid(row=2, column=1, padx=5, pady=5)

        ttk.Button(self.root, text="Login", command=self.login).grid(row=2, column=0, pady=10)

        self.update_login_fields()

    def update_login_fields(self):
        if self.login_type.get() == "teacher":
            self.label_username.grid()
            self.entry_username.grid()
            self.label_student_id.grid_remove()
            self.entry_student_id.grid_remove()
        else:
            self.label_username.grid_remove()
            self.entry_username.grid_remove()
            self.label_student_id.grid()
            self.entry_student_id.grid()

    def login(self):
        login_type = self.login_type.get()
        password = self.entry_password.get()

        if login_type == "teacher":
            username = self.entry_username.get()
            if username == "teacher" and password == "teacher123":
                self.root.destroy()
                root = tk.Tk()
                app = TeacherApp(root)
                root.mainloop()
            else:
                messagebox.showerror("Login Failed", "Invalid username or password")
        else:
            student_id = self.entry_student_id.get()
            if student_id and password:
                students = read_csv(STUDENT_FILE)
                student_found = False
                for student in students[1:]:
                    if student[0] == student_id:
                        if verify_password(password, student[6]):
                            student_found = True
                            self.root.destroy()
                            root = tk.Tk()
                            app = StudentApp(root, student[0])
                            root.mainloop()
                            break
                if not student_found:
                    messagebox.showerror("Login Failed", "Invalid Student ID or password")
            else:
                messagebox.showwarning("Input Error", "Please enter Student ID and Password")

# Main Application
if __name__ == "__main__":
    root = tk.Tk()
    app = LoginApp(root)
    root.mainloop()